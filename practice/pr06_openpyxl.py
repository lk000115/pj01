from openpyxl import load_workbook
wb = load_workbook(r'F:\books\工资条.xlsx')
ws = wb.worksheets[1]
class FooError(ValueError):
    pass
print(ws)
lt = []
st = []
for cell in ws[1]:
    st.append(cell.value)
# print(st)
if st[0] != '工号':
    raise FooError('格式错误')

for r in ws.iter_rows():
    lt = []
    s1 = ''
    for k in r:
        lt.append(k.value)
    for i in range(len(lt)):
        s1 = s1 + st[i] + ' : ' + str(lt[i]) + '\n'
    print(f'{s1}', lt[0] )







