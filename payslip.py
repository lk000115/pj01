from openpyxl import load_workbook
from wechat_work import WechatWork
import tkinter as tk
from tkinter import filedialog


corpid = 'wx86ade657d9d15fc5'
appid = '1000015'
corpsecret = 'vP_tz5m43cgFsQ-phNeWgZBhIelYxv3OEaXB4JLrHHY'
users = []
w = WechatWork(corpid=corpid,
               appid=appid,
               corpsecret=corpsecret)

root = tk.Tk()
root.title('工资条发送小工具')
root.geometry('400x400')
var = tk.StringVar()
var2 = tk.StringVar()

def get_file():
    global Filepath
    sub_root = tk.Toplevel(root)
    sub_root.withdraw()
    Filepath = filedialog.askopenfilename()
    var.set(Filepath)
    sub_root.destroy()  # 销毁窗体


def send_pay(x=0):
    wb = load_workbook(Filepath)
    sheet = wb.worksheets[x]
    st = []
    for cell in sheet[1]:
        st.append(cell.value)
    i = 0
    for row in sheet.iter_rows(min_row=2):  # 从第二行开始读,循环获取表中的每行值
        lt = []
        users = []
        gz_str = ''
        for r in row:                       # 循环行获取单元格的值
            lt.append(r.value)
        for k in range(len(lt)):           # 把行列表转换为字符串
            try:
                gz_str = gz_str + st[k] + ' : ' + str(lt[k]) + '\n'
            except TypeError as e:
                var2.set(f'发送失败,表数据格式错误')
                raise
        users.append(lt[1])
        print(f'{gz_str}',users)
        # if len(lt) > 40:
        #     var2.set('表格格式错误')
        # elif st[1] != '工号':
        #     var2.set('第二列必须是工号')
        # else:
        #     i += 1
        #     w.send_text(f'{gz_str}',users)
        #     var2.set(f'发送成功,总共发送{i}条数据')

def send_pay1():
    send_pay(1)


l = tk.Label(root, textvariable=var, font=('Arial', 12), width=60, height=6)
l.pack()

btn = tk.Button(root, text="添加工资条", command=get_file)
btn.pack()

l2 = tk.Label(root, textvariable=var2, font=('Arial', 12), width=30, height=6)
l2.pack()

btn2 = tk.Button(root, text="发送管理工资条", command=send_pay)
btn2.pack()

btn3 = tk.Button(root, text="发送计时工资条", command=send_pay1)
btn3.pack()

root.mainloop()



